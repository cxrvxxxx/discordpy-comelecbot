import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as column

workfile_path = r'./data/workfile'

class Workfile:
    def __init__(self, fn):
        self.fn = fn
        self.wb = load_workbook(f'{workfile_path}/{self.fn}')
        self.ws = self.wb.active

    def fetch_row(self, row_id: int, col_end: int, col_start: int  = 1):
        fields = []

        for c in range(col_start, col_end):
            fields.append(self.ws[f'{column(c)}{row_id + 1}'].value)

        return fields

    def delete(self):
        os.remove(fr'{workfile_path}/{self.fn}')

    @staticmethod
    def check_exists(fn):
        if fn in os.listdir(workfile_path): return True
        return False

    @staticmethod
    def get_workfile():
        if len(os.listdir(workfile_path)) > 0:
            return os.listdir(workfile_path)[0]

    def cell(self, col: int, row: int):
        return self.ws[f'{column(col)}{row}']
