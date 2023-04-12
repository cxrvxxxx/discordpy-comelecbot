import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter as column

from . import workfile_path

class Workfile:
    def __init__(self, fn):
        self.fn = fn
        self.wb = load_workbook(f'{workfile_path}/{self.fn}')
        self.ws = self.wb.active

    def fetch_row(self, row_id):
        fields = []

        c = 1
        while True:
            value = self.ws[f'{column(c)}{row_id + 1}'].value
            if not value: break
            fields.append(value)
            c += 1

        return fields

    def delete(self):
        os.remove(fr'{workfile_path}/{self.fn}')

    @staticmethod
    def check_exists(fn):
        if fn in os.listdir(workfile_path): return True
        return False

    @staticmethod
    def get_workfile():
        if len(os.listdir(workfile_path)) > 0:
            return os.listdir(workfile_path)[0]

    def cell(self, col: int, row: int):
        return self.ws[f'{column(col)}{row}']
